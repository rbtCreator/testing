#!/usr/bin/env python3
# -*- coding: utf-8 -*-


#import necessary libraries for initiating Qt_object
from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.Qt import QTimer
from PyQt5.QtCore import Qt, pyqtSignal, QObject, QByteArray, QEvent

import csv
import os
import numpy as np
import openpyxl as opx
import math


import init
import quest
import result 



class StartWindow(QtWidgets.QWidget, init.Ui_Form):
    def __init__(self, nextStepFunc, dataSet, img):
        super(StartWindow, self).__init__()
        self.setupUi(self)
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)

        self.nextStepFunc = nextStepFunc
        self.dataSet = dataSet
        self.img = img

        self.start_but.clicked.connect(self.buttHandler)

        self.VAS.setText("")

        self.time = None
        self.sampleCount = None

        self.loadTestData()
        self.localSetupUi()
        
    def loadTestData(self):
        path = os.path.abspath(os.curdir)
        path = os.path.split(path)[0]
        source = os.path.join(path, "testData", "source.csv")
        
        with open(source, newline='') as csvfile:
            reader = csv.DictReader(csvfile)
            rows = list(reader)
            ex = rows[0]
            self.time = int(ex["time"])
            self.test_name.setText(ex["name"])
            self.sampleCount = len(rows)
            self.dataSet(rows)
    
        
    def localSetupUi(self):
        self.perNumbC.setInputMask("AA")
        if self.time == -1:
            self.time_limit.setText("Время выполнения: неограниченно")
        else:
            self.time_limit.setText("Время выполнения: {} мин.".format(str(self.time)))
        self.q_numb.setText("Кол-во вопросов: {}".format(self.sampleCount))
        
        self.palette = QtGui.QPalette()
        self.palette.setBrush(QtGui.QPalette.Window, QtGui.QBrush(self.img))
        self.setPalette(self.palette)
        

    def buttHandler(self):
        isCorrectData = True
        number = ""
        name = self.fio_edit.text()
        if len(name.split()) != 3:
            self.fio_edit.setText("")
            self.fioErrorLabel.setText("Фамилия Имя Отчество")
            isCorrectData = False

        if isCorrectData:
            didgNum = self.perNumbD.text()
            if didgNum.isdigit():
                number = self.perNumbC.text() + "_" + didgNum
            elif didgNum == "ГП РФ":
                number = "no_nummber"
            else:
                self.perNumbC.setText("")
                self.perNumbD.setText("")
                self.numbErrorLabel.setText("НР-493890  или ГП РФ")
                isCorrectData = False

        if isCorrectData:
            self.nextStepFunc(name, number)




class ResultWin(QtWidgets.QWidget, result.Ui_Form):
    def __init__(self, setPhr, img):
        super(ResultWin, self).__init__()
        self.setupUi(self)
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)

        self.setPhr = setPhr
        self.img = img

        self.butt.clicked.connect(self.btnHandler)

        self.VAS.setText("")

        self.adv = None
        self.ans = None
        self.userChoice = None
        self.name = None
        self.number = None
        self.res = None
        self.docName = None

        self.localSetupUi()
        self.loadAdvices()

    def setDocName(self, name):
        self.docName = name

    def setNameNumber(self, name, number):
        self.name = name
        self.number = number

    def xlsxOutput(self):
        path = os.path.abspath(os.curdir)
        path = os.path.split(path)[0]
        path = os.path.join(path, "testData", "result", self.docName)
        workbook = opx.load_workbook(path)
        worksheet = workbook.worksheets[0]
        last_row = worksheet.max_row
        if not worksheet.cell(row=last_row, column=1).value is None:
            last_row += 1

        worksheet.cell(row=last_row, column=1).value = self.name
        worksheet.cell(row=last_row, column=2).value = self.number
        worksheet.cell(row=last_row, column=3).value = self.res
        workbook.save(path)
        

    def loadAdvices(self):
        path = os.path.abspath(os.curdir)
        path = os.path.split(path)[0]
        with open(os.path.join(path, "testData", "adv.csv"), newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            adv, phr = [], []
            for row in reader:
                if row["adv"] != "-":
                    st = row["adv"]
                    st = st.replace("/", ",")
                    adv.append(st)
                if row["phr"] != "-":
                    st = row["phr"]
                    st = st.replace("/", ",")
                    phr.append(row["phr"])
        
        self.adv = adv
        self.setPhr(phr)

    def setUserChoice(self, choice):
        self.decryptChoice()
        self.userChoice = choice
        match = 0
        for step, choice in enumerate(self.userChoice):
            if choice == self.ans[step]:
                match += 1

        per = int(match/len(self.ans)*100)
        self.res = per
        self.res_per.setText("Решено: {}%  ({} из {})".format(per, match, len(self.ans)))
        #print(per, match, len(self.ans))
        per -= 55
        per = per // 5
        if per < 0: per = 0
        #print("make {} matches and {} per".format(match, per))
        #print(self.adv[per])
        self.advice.setText(self.adv[per])

    def decryptChoice(self):
        for i, el in enumerate(self.ans):
            self.ans[i] = int(math.sqrt(el + 121) - 11)

    def btnHandler(self):
        self.xlsxOutput()
        self.close()
        

    def setData(self, data):
        ans = []
        for row in data:
            ans.append(int(row["ans"]))

        self.ans = ans

    def localSetupUi(self):
        self.palette = QtGui.QPalette()
        self.palette.setBrush(QtGui.QPalette.Window, QtGui.QBrush(self.img))
        self.setPalette(self.palette)

    def closeEvent(self, event):
        reply = QtWidgets.QMessageBox.question(self, 'Message', "Завершить выполнение?", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.No)
        if reply == QtWidgets.QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()
        
               
        





class QuestForm(QtWidgets.QWidget, quest.Ui_Form):
    def __init__(self, nextStepFunc, img):
        super(QuestForm, self).__init__()
        self.setupUi(self)
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)

        self.nextStepFunc = nextStepFunc
        self.img = img

        self.first_ans.toggled.connect(self.onSelect)
        self.first_ans.name = 0 #"first_ans"
        self.second_ans.toggled.connect(self.onSelect)
        self.second_ans.name = 1 #"second_ans"
        self.third_ans.toggled.connect(self.onSelect)
        self.third_ans.name = 2 #"third_ans"
        self.fourth_ans.toggled.connect(self.onSelect)
        self.fourth_ans.name = 3 #"fourth_ans

        self.btn_group = QtWidgets.QButtonGroup()
        self.btn_group.addButton(self.first_ans)
        self.btn_group.addButton(self.second_ans)
        self.btn_group.addButton(self.third_ans)
        self.btn_group.addButton(self.fourth_ans)
        self.pushButton.clicked.connect(self.nextStep)
        self.pushButton_2.clicked.connect(self.prevStep)

        self.VAS.setText("")

        self.Qtimer = None
        self.testData = []
        self.choice = None
        self.phrases = None
        self.step = 0
        self.pitchSize = None
        self.time = None
        self.stopCount = None
        self.currTime = 0

        self.not_finished = True
        

        self.localSetupUi()

    def timeOut(self):
        if -1 in self.choice:
            print("OMG")
            ans = []
            for sample in self.testData:
                ans.append(int(sample["ans"]))
            for i, el in enumerate(self.choice):
                if el == -1:
                    self.choice[i] = 0
                    if ans[i] == 0: self.choice[i] = 1

        self.nextStepFunc(self.choice)

    def setupTimer(self):
        if self.time > 0:
            self.stopCount = self.time*60

    def timerStart(self):
        self.timer.setText("0:00")
        self.Qtimer = QTimer(self)
        self.Qtimer.timeout.connect(self.loop)
        self.Qtimer.start(1000)

    def loop(self):
        self.currTime += 1
        if self.not_finished:
            if not self.stopCount is None:
                if self.currTime == self.stopCount:
                    self.Qtimer.stop()
                    self.timeOut()
                if self.stopCount - self.currTime < 30: self.timer.setStyleSheet("color: rgb(219, 35, 35);")

            time = str(self.currTime // 60) + ":"
            if self.currTime % 60 < 10: time += "0" + str(self.currTime % 60)
            else: time += str(self.currTime % 60)
            self.timer.setText(time)

            
        else:
            self.Qtimer.stop()
        

    def setPhrases(self, phr):
        self.phrases = phr

    def setData(self, data):
        self.pitchSize = len(data)
        self.choice = [ -1 for el in range(self.pitchSize)]

        dataOrder = np.random.permutation(self.pitchSize)
        for ind in dataOrder:
            self.testData.append(data[ind])

        setup = self.testData[0]
        self.time = int(setup["time"])
        self.test_name.setText(setup["name"])
        self.setupTimer()

        self.guiUpdate(step=0)

    def nextStep(self):
        if self.step == self.pitchSize-1:
            self.not_finished = False
            self.nextStepFunc(self.choice)
        else:
            self.step += 1
            self.guiUpdate(step=self.step)

    def prevStep(self):
        self.step -= 1
        self.guiUpdate(step=self.step)

    def onSelect(self):
        radioBtn = self.sender()
        if radioBtn.isChecked():
            self.choice[self.step] = radioBtn.name

        if self.step == self.pitchSize-1:
            if not -1 in self.choice:
                self.pushButton.setEnabled(True)


    def guiUpdate(self, step=None):
        sample = self.testData[step]

        self.quest.setText(sample["text"])
        self.phrase.setText(np.random.choice(self.phrases))
        self.first_ans.setText(sample["q_1"])
        self.second_ans.setText(sample["q_2"])
        self.third_ans.setText(sample["q_3"])
        self.fourth_ans.setText(sample["q_4"])

        self.btn_group.setExclusive(False) # переводим группу из екслюзивного ржима, в нем возможен множественный выбор/не выбрать ни одной
        for radio in self.btn_group.buttons():
            radio.setChecked(False)  # снимаем отметки
        self.btn_group.setExclusive(True)

        if self.choice[step] != -1:
            buttons = self.btn_group.buttons()
            button = buttons[self.choice[step]]
            button.setChecked(True)


        self.pushButton.setEnabled(True)
        self.pushButton_2.setEnabled(True)
        self.pushButton.setText("следующий")
        
        if step == 0:
            self.pushButton_2.setEnabled(False)
        elif step == self.pitchSize-1:
            self.pushButton.setText("завершить")
            if -1 in self.choice:
                self.pushButton.setEnabled(False)
            

        self.quest_numb.setText("Вопрос №" + str(self.step + 1))

    def localSetupUi(self):
        self.palette = QtGui.QPalette()
        self.palette.setBrush(QtGui.QPalette.Window, QtGui.QBrush(self.img))
        self.setPalette(self.palette)
        
    

        
